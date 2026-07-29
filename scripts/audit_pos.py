#!/usr/bin/env python3
"""
Audit PPS purchase orders.

Checks:
  1. PO number in the file name vs the PO number inside the sheet
  2. Lines annotated as excluded that are still summed into the total
  3. Whether every PO line can be joined to a master SKU
  4. Stale footer metadata (e.g. "SKUs: 21" on a 38-line PO)

Usage: python3 scripts/audit_pos.py
"""

import csv
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "source")

PO_FILES = [
    dict(file="PO_film.xlsx", label="Stick Film", total_row=76,
         filename_po="41626", footer_rows=(83, 89)),
    dict(file="PO_protein_pouches.xlsx", label="Protein Pouch", total_row=76,
         filename_po="42926", footer_rows=(83, 88)),
    dict(file="PO_pancake_etc_pouches.xlsx", label="Mix Pouch", total_row=48,
         filename_po="43026", footer_rows=(54, 58)),
]

ALIASES = {
    "raspberrycheescake": "raspberrycheesecake",
    "cafemocha": "cafmocha",
    "snickerdoodlecookie": "snickerdoodle",
    "gfflour": "glutenfreeflour",
}


def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def build_index():
    path = os.path.join(SRC, "master_artwork_sheet.csv")
    idx = {}
    for r in csv.DictReader(open(path)):
        f = r["flavor"]
        line = ("WHEY" if "Whey" in f else "BEEF" if "Beef" in f
                else "PLANT" if "Plant" in f else "OTHER")
        flav = re.sub(r"\s*(Whey|Beef|Plant)\s*Protein\s*(Pouch|Stick)$", "", f)
        flav = re.sub(
            r"\s*(Protein\s*)?(Pancake Mix|Crepe Mix|Cupcake Mix|Donut Mix|"
            r"Oatmeal Cup|Pouch|Stick|Mix|Cup)$", "", flav).strip()
        idx[(line, r["packaging type"].upper(), norm(flav))] = r["sku"]
    return idx


def main():
    idx = build_index()
    total_over_u = total_over_d = 0.0
    problems = 0

    for cfg in PO_FILES:
        ws = openpyxl.load_workbook(os.path.join(SRC, cfg["file"]), data_only=True).active
        po = str(ws["F15"].value or "").lstrip("#").strip()

        print(f"\n{'=' * 68}\n{cfg['label']}  —  {cfg['file']}")

        if cfg["filename_po"] not in po:
            print(f"  [HIGH] PO number mismatch: file name says {cfg['filename_po']}, "
                  f"sheet says {po}")
            problems += 1
        else:
            print(f"  PO number {po} — file name and sheet agree")

        lines, flagged, unjoined = [], [], []
        for row in range(28, cfg["total_row"]):
            d, f_, q = ws[f"B{row}"].value, ws[f"C{row}"].value, ws[f"E{row}"].value
            c, ext, note = ws[f"G{row}"].value, ws[f"H{row}"].value, ws[f"I{row}"].value
            if not d or not q:
                continue
            lines.append((d, f_, q, c, ext, note))
            if note:
                flagged.append((d, f_, q, ext, note))
            du = d.upper()
            ln = ("WHEY" if "WHEY" in du else "BEEF" if "BEEF" in du
                  else "PLANT" if "PLANT" in du else "OTHER")
            fmt = "STICK" if ("STICK" in du or "FILM" in du) else "POUCH"
            nf = ALIASES.get(norm(f_), norm(f_))
            if (ln, fmt, nf) not in idx and not any(
                    (o, f2, nf) in idx for o in ("OTHER",) for f2 in ("POUCH", "BOX", "CUP")):
                unjoined.append((d, str(f_).strip()))

        tq = sum(l[2] for l in lines)
        tt = sum(l[4] for l in lines)
        print(f"  {len(lines)} lines | total as written {tq:,.0f} u / ${tt:,.2f}")

        if flagged:
            fq = sum(f[2] for f in flagged)
            ft = sum(f[3] for f in flagged)
            total_over_u += fq
            total_over_d += ft
            problems += len(flagged)
            print(f"  [CRITICAL] {len(flagged)} lines annotated as excluded are still "
                  f"inside the SUM:")
            for d, f_, q, ext, note in flagged:
                print(f"      {d:<32}{str(f_).strip():<18}{q:>9,.0f} u  "
                      f"${ext:>10,.2f}   \"{note}\"")
            print(f"      overstated by {fq:,.0f} u / ${ft:,.2f}")
            print(f"      corrected total: {tq - fq:,.0f} u / ${tt - ft:,.2f}")

        if unjoined:
            problems += len(unjoined)
            print(f"  [CRITICAL] {len(unjoined)} lines cannot be joined to a SKU:")
            for d, f_ in unjoined:
                print(f"      '{f_}'  ({d})")
        else:
            print(f"  all {len(lines)} lines join to a master SKU "
                  f"(after alias repair)")

        lo, hi = cfg["footer_rows"]
        for r in range(lo, hi + 1):
            v = ws[f"B{r}"].value
            if isinstance(v, str) and v.startswith("SKUs:"):
                claimed = int(re.search(r"\d+", v).group())
                if claimed != len(lines):
                    print(f"  [HIGH] stale footer: says '{v}' but the PO has "
                          f"{len(lines)} line items")
                    problems += 1
            if isinstance(v, str) and "?" in v:
                print(f"  [MEDIUM] footer sent to vendor contains a placeholder: '{v}'")

    print(f"\n{'=' * 68}")
    print(f"COMBINED OVERSTATEMENT: {total_over_u:,.0f} units / ${total_over_d:,.2f}")
    print(f"Total problems: {problems}")
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
