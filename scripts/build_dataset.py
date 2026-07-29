#!/usr/bin/env python3
"""
Normalize the legacy ProDough spreadsheets into a relational dataset that can be
imported directly into Airtable (or any database).

Inputs  : source/master_artwork_sheet.csv, source/PO_*.xlsx
Outputs : data/*.csv

Every output table keys off `sku`, which is the join key the current
spreadsheets do not have.

Run:  python3 scripts/build_dataset.py
"""

import csv
import os
import re
from collections import OrderedDict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "source")
OUT = os.path.join(ROOT, "data")

# --------------------------------------------------------------------------
# Packaging specs: the master sheet repeats the same 3 structures on 118 rows.
# Normalize them into 5 spec records that SKUs point at.
# --------------------------------------------------------------------------
SPECS = [
    {
        "spec_id": "SPEC-POUCH-LG",
        "spec_name": "Large Protein Pouch (10x10+3G)",
        "format": "Pouch",
        "material_structure": "Karess Soft Touch Matte PET / White 48ga METPET / 3.0mil Clear LLDPE",
        "zipper": "Aplix",
        "print_process": "CMYK",
        "trim_length_mm": "254",
        "trim_width_mm": "254",
        "gusset_mm": "76.2",
        "front_panel_mm": "",
        "wind_direction": "",
        "core_in": "",
        "vendor": "Professional Packaging Services (PPS)",
        "last_unit_cost": "0.506",
        "dieline_required": "TRUE",
        "vendor_spec_string": "Size: 10.00 W x 10.00 L + 3.00 G",
        "notes": "Used by whey / beef / plant protein pouches.",
    },
    {
        "spec_id": "SPEC-POUCH-SM",
        "spec_name": "Small Mix Pouch (8x9+3G)",
        "format": "Pouch",
        "material_structure": "Karess Soft Touch Matte PET / White 48ga METPET / 3.0mil Clear LLDPE",
        "zipper": "Aplix",
        "print_process": "CMYK",
        "trim_length_mm": "228.6",
        "trim_width_mm": "203.2",
        "gusset_mm": "76.2",
        "front_panel_mm": "",
        "wind_direction": "",
        "core_in": "",
        "vendor": "Professional Packaging Services (PPS)",
        "last_unit_cost": "0.434",
        "dieline_required": "TRUE",
        "vendor_spec_string": "Size: 8.00 W x 9.00 L + 3.00 G",
        "notes": "Pancake / crepe / cupcake / GF flour pouches.",
    },
    {
        "spec_id": "SPEC-STICK-LG",
        "spec_name": "Large Stick Film (175mm web)",
        "format": "Stick",
        "material_structure": "#781 PET Soft Touch on #858 COS Web metallocene",
        "zipper": "",
        "print_process": "CMYK",
        "trim_length_mm": "175",
        "trim_width_mm": "140",
        "gusset_mm": "",
        "front_panel_mm": "66",
        "wind_direction": "3",
        "core_in": "3.00",
        "vendor": "Professional Packaging Services (PPS)",
        "last_unit_cost": "0.0749",
        "dieline_required": "TRUE",
        "vendor_spec_string": "Size: 6.8898 Web x 5.5118 | Structure: .50ga Soft Touch Matte PET/280ga White Stickweb | Core 3.00 | Max OD: ?",
        "notes": "CONFLICT: master sheet and PO footer describe this structure differently. Reconcile with PPS and keep one string.",
    },
    {
        "spec_id": "SPEC-BOX-DONUT",
        "spec_name": "Donut Mix Carton",
        "format": "Box",
        "material_structure": "",
        "zipper": "",
        "print_process": "",
        "trim_length_mm": "",
        "trim_width_mm": "",
        "gusset_mm": "",
        "front_panel_mm": "",
        "wind_direction": "",
        "core_in": "",
        "vendor": "",
        "last_unit_cost": "",
        "dieline_required": "",
        "vendor_spec_string": "",
        "notes": "GAP: no physical spec recorded for 26 donut mix SKUs.",
    },
    {
        "spec_id": "SPEC-CUP-OAT",
        "spec_name": "Oatmeal Cup",
        "format": "Cup",
        "material_structure": "",
        "zipper": "",
        "print_process": "",
        "trim_length_mm": "",
        "trim_width_mm": "",
        "gusset_mm": "",
        "front_panel_mm": "",
        "wind_direction": "",
        "core_in": "",
        "vendor": "",
        "last_unit_cost": "",
        "dieline_required": "",
        "vendor_spec_string": "",
        "notes": "GAP: no physical spec recorded for 3 oatmeal cup SKUs.",
    },
]

SPEC_BY_FORMAT_AND_SIZE = {
    ("Pouch", "254"): "SPEC-POUCH-LG",
    ("Pouch", "228.6"): "SPEC-POUCH-SM",
    ("Stick", "175"): "SPEC-STICK-LG",
    ("Box", ""): "SPEC-BOX-DONUT",
    ("Cup", ""): "SPEC-CUP-OAT",
}

LINE_RULES = [
    ("Whey Protein", "Whey Protein"),
    ("Beef Protein", "Beef Protein"),
    ("Plant Protein", "Plant Protein"),
    ("Protein Pancake Mix", "Pancake Mix"),
    ("Protein Crepe Mix", "Crepe Mix"),
    ("Protein Cupcake Mix", "Cupcake Mix"),
    ("Protein Donut Mix", "Donut Mix"),
    ("Protein Oatmeal Cup", "Oatmeal Cup"),
    ("Daily Recharge", "Daily Recharge"),
    ("Gluten Free Flour", "Gluten Free Flour"),
]

HEX_RE = re.compile(r"^(HEX|HX)\s+([0-9A-Fa-f]{6})$")
PMS_RE = re.compile(r"^PMS\s+\S")


def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def gtin_valid(g):
    g = (g or "").strip()
    if not g.isdigit() or len(g) not in (8, 12, 13, 14):
        return False
    total = 0
    for i, d in enumerate(reversed(g[:-1])):
        total += int(d) * (3 if i % 2 == 0 else 1)
    return (10 - total % 10) % 10 == int(g[-1])


def product_line(flavor):
    for needle, label in LINE_RULES:
        if needle in flavor:
            return label
    return "Misc"


def base_flavor(flavor):
    f = flavor
    f = re.sub(r"\s*(Whey|Beef|Plant)\s*Protein\s*(Pouch|Stick)$", "", f)
    f = re.sub(r"\s*Protein\s*(Pancake Mix|Crepe Mix|Cupcake Mix|Donut Mix|Oatmeal Cup)$", "", f)
    f = re.sub(r"\s*(Pouch|Stick|Mix|Cup)$", "", f)
    return f.strip()


APP_DATA = os.path.join(ROOT, "app", "data")


def write(name, fieldnames, rows):
    # Written twice: data/ is the canonical dataset, app/data/ is the copy the
    # deployed seed reads (Railway builds with app/ as its root).
    targets = [OUT] + ([APP_DATA] if os.path.isdir(APP_DATA) else [])
    for target in targets:
        with open(os.path.join(target, name), "w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
    print(f"  wrote {name:<28} {len(rows):>4} rows")


# --------------------------------------------------------------------------
def build_products():
    src = os.path.join(SRC, "master_artwork_sheet.csv")
    rows = list(csv.DictReader(open(src)))

    products, colors, issues = [], [], []

    for r in rows:
        flavor = r["flavor"].strip()
        sku = r["sku"].strip()
        gtin = r["gtin/barcode"].strip()
        fmt = r["packaging type"].strip()
        trim = r["trim length"].strip().replace("mm", "")
        spec = SPEC_BY_FORMAT_AND_SIZE.get((fmt, trim), "")

        sku_is_placeholder = sku.isdigit()
        if sku_is_placeholder:
            issues.append(
                dict(
                    severity="High",
                    area="SKU",
                    sku=sku,
                    flavor=flavor,
                    issue="SKU column holds a Shopify variant ID, not a SKU code",
                    detail=f"value '{sku}' is 14 digits; needs a real code (suggest HBF-* to match the beef sticks)",
                )
            )
        if not gtin_valid(gtin):
            issues.append(
                dict(severity="Critical", area="GTIN", sku=sku, flavor=flavor,
                     issue="GTIN fails GS1 check digit", detail=gtin)
            )
        if not spec:
            issues.append(
                dict(severity="Medium", area="Spec", sku=sku, flavor=flavor,
                     issue="No packaging spec resolves for this SKU",
                     detail=f"format={fmt or '(blank)'} trim={trim or '(blank)'}")
            )
        if fmt == "Stick" and not r["eye mark color"].strip():
            issues.append(
                dict(severity="High", area="Artwork", sku=sku, flavor=flavor,
                     issue="Stick pack has no eyemark colour specified",
                     detail="bagger photo-eye cannot register without it")
            )

        pms_tokens = [t.strip() for t in (r["pms spot colors"] or "").split(",") if t.strip()]
        hex_tokens = [t.strip() for t in (r["hex spot colors"] or "").split(",") if t.strip()]
        if not pms_tokens:
            issues.append(
                dict(severity="Medium", area="Colour", sku=sku, flavor=flavor,
                     issue="No brand colours recorded", detail="")
            )
        for i in range(max(len(pms_tokens), len(hex_tokens))):
            pms = pms_tokens[i] if i < len(pms_tokens) else ""
            hx = hex_tokens[i] if i < len(hex_tokens) else ""
            hex_ok = bool(HEX_RE.match(hx)) if hx else False
            pms_ok = bool(PMS_RE.match(pms)) and pms != "PMS --" if pms else False
            colors.append(
                dict(sku=sku, slot=i + 1, pms=pms, hex=hx,
                     hex_valid="TRUE" if hex_ok else "FALSE",
                     pms_valid="TRUE" if pms_ok else "FALSE")
            )
            if hx and not hex_ok:
                issues.append(
                    dict(severity="High", area="Colour", sku=sku, flavor=flavor,
                         issue="Invalid hex value — cannot be rendered",
                         detail=f"slot {i+1}: '{hx}'")
                )
            if pms and not pms_ok:
                issues.append(
                    dict(severity="Low", area="Colour", sku=sku, flavor=flavor,
                         issue="Colour slot is not a valid PMS reference",
                         detail=f"slot {i+1}: '{pms}'")
                )

        products.append(
            OrderedDict(
                sku=sku,
                gtin=gtin,
                gtin_valid="TRUE" if gtin_valid(gtin) else "FALSE",
                product_line=product_line(flavor),
                format=fmt,
                flavor=flavor,
                base_flavor=base_flavor(flavor),
                spec_id=spec,
                eyemark_color=r["eye mark color"].strip(),
                dieline_required=r["die line required"].strip(),
                status="Active",
                shopify_sku="",
                shopify_variant_id=sku if sku_is_placeholder else "",
                mrp_formula_id="",
                formula_rev="",
                artwork_version="",
                artwork_status="Unknown",
                nfp_version="",
                nfp_last_reviewed="",
                data_quality_flag="NEEDS REVIEW" if sku_is_placeholder else "",
                notes="",
            )
        )

    # SKU stem collisions
    stems = {}
    for p in products:
        m = re.match(r"^([A-Z]+-[A-Z]+)", p["sku"])
        if m:
            stems.setdefault(m.group(1), []).append(p)
    for stem, group in stems.items():
        distinct = {g["base_flavor"] for g in group}
        if len(distinct) > 1:
            for g in group:
                issues.append(
                    dict(severity="High", area="SKU", sku=g["sku"], flavor=g["flavor"],
                         issue=f"SKU stem '{stem}' is reused across different flavours",
                         detail="; ".join(sorted(f"{x['sku']}={x['base_flavor']}" for x in group)))
                )

    return products, colors, issues


# --------------------------------------------------------------------------
PO_FILES = [
    dict(file="PO_film.xlsx", category="Stick Film", total_row=76,
         default_spec="SPEC-STICK-LG", filename_po="41626"),
    dict(file="PO_protein_pouches.xlsx", category="Protein Pouch", total_row=76,
         default_spec="SPEC-POUCH-LG", filename_po="42926"),
    dict(file="PO_pancake_etc_pouches.xlsx", category="Mix Pouch", total_row=48,
         default_spec="SPEC-POUCH-SM", filename_po="43026"),
]

# manual repairs for PO flavour text that does not match the master sheet
# The master sheet writes "Café Mocha" with an accent; norm() strips the é and
# yields "cafmocha", while the PO writes plain "Cafe Mocha" -> "cafemocha".
# A single accented character silently breaks the join. This is exactly the
# class of failure the new system removes by joining on SKU instead of on text.
ALIASES = {
    "raspberrycheescake": "raspberrycheesecake",
    "cafemocha": "cafmocha",
    "snickerdoodlecookie": "snickerdoodle",
    "gfflour": "glutenfreeflour",
}


def po_line_to_product_line(description):
    """Map a PO line's free-text description onto a product_line.

    Keying on flavour alone is not enough: 'Chocolate' exists as a pancake mix,
    a crepe mix and a cupcake mix, all in the same POUCH format. Without the
    product type in the key they collide and the later one silently wins.
    """
    d = description.upper()
    if 'WHEY' in d:
        return 'Whey Protein'
    if 'BEEF' in d:
        return 'Beef Protein'
    if 'PLANT' in d:
        return 'Plant Protein'
    if 'PANCAKE' in d:
        return 'Pancake Mix'
    if 'CREPE' in d:
        return 'Crepe Mix'
    if 'CUPCAKE' in d:
        return 'Cupcake Mix'
    if 'DONUT' in d:
        return 'Donut Mix'
    if 'OATMEAL' in d:
        return 'Oatmeal Cup'
    if 'GF FLOUR' in d or 'GLUTEN' in d:
        return 'Gluten Free Flour'
    if 'DAILY RECHARGE' in d or 'RECHARGE' in d:
        return 'Daily Recharge'
    return None


def build_pos(products):
    # Keyed on (product_line, format, base_flavor) — the product type has to be
    # part of the key or pancake/crepe/cupcake "Chocolate" all collapse together.
    index = {}
    for p in products:
        index[(p["product_line"], p["format"].upper(), norm(p["base_flavor"]))] = p["sku"]

    pos, lines, issues = [], [], []

    for cfg in PO_FILES:
        ws = openpyxl.load_workbook(os.path.join(SRC, cfg["file"]), data_only=True).active
        po_number = str(ws["F15"].value or "").lstrip("#").strip()
        po_date = ws["F12"].value
        po_date = po_date.strftime("%Y-%m-%d") if hasattr(po_date, "strftime") else str(po_date)

        if cfg["filename_po"] not in po_number:
            issues.append(
                dict(severity="High", area="PO", sku="", flavor="",
                     issue="PO number in the file name does not match the number inside the sheet",
                     detail=f"{cfg['file']} -> filename says {cfg['filename_po']}, sheet says {po_number}")
            )

        n = 0
        excl_qty = excl_cost = 0.0
        tot_qty = tot_cost = 0.0

        for row in range(28, cfg["total_row"]):
            desc = ws[f"B{row}"].value
            flav = ws[f"C{row}"].value
            qty = ws[f"E{row}"].value
            cost = ws[f"G{row}"].value
            ext = ws[f"H{row}"].value
            note = ws[f"I{row}"].value
            if not desc or not qty:
                continue
            n += 1
            du = desc.upper()
            prod_line = po_line_to_product_line(desc)
            fmt = "STICK" if ("STICK" in du or "FILM" in du) else "POUCH"
            nf = norm(flav)
            nf = ALIASES.get(nf, nf)
            sku = index.get((prod_line, fmt, nf), "") if prod_line else ""
            if not sku and prod_line:
                # single-SKU lines (GF flour) name the product, not a flavour
                candidates = [
                    p for p in products
                    if p["product_line"] == prod_line and p["format"].upper() == fmt
                ]
                if len(candidates) == 1:
                    sku = candidates[0]["sku"]
            if not sku:
                issues.append(
                    dict(severity="Critical", area="PO", sku="", flavor=str(flav).strip(),
                         issue="PO line cannot be matched to any SKU",
                         detail=f"PO {po_number} — '{str(flav).strip()}' ({desc})")
                )

            excluded = "TRUE" if note else "FALSE"
            if note:
                excl_qty += qty
                excl_cost += ext
                issues.append(
                    dict(severity="Critical", area="PO", sku=sku, flavor=str(flav).strip(),
                         issue="Line is marked as excluded but is still counted in the PO total",
                         detail=f"PO {po_number} — {qty:,.0f} u / ${ext:,.2f} — note: \"{note}\"")
                )
            tot_qty += qty
            tot_cost += ext

            lines.append(
                OrderedDict(
                    po_number=po_number, line_no=n, sku=sku,
                    description=desc, flavor=str(flav).strip(),
                    spec_id=cfg["default_spec"],
                    qty=int(qty), unit_cost=cost, ext_cost=round(ext, 2),
                    excluded=excluded, exclusion_note=note or "",
                    line_status="Ordered", expected_date="",
                    qty_received="", date_received="", qc_status="", location="",
                )
            )

        pos.append(
            OrderedDict(
                po_number=po_number, po_date=po_date,
                vendor="Professional Packaging Services (PPS)",
                vendor_contact="Michael Rino <mrino@propackserv.com>",
                category=cfg["category"], source_file=cfg["file"],
                line_count=n,
                total_qty_as_written=int(tot_qty),
                total_cost_as_written=round(tot_cost, 2),
                excluded_qty=int(excl_qty),
                excluded_cost=round(excl_cost, 2),
                total_qty_corrected=int(tot_qty - excl_qty),
                total_cost_corrected=round(tot_cost - excl_cost, 2),
                status="Open", authorized_by="Daniel Augustyn, CEO",
                notes="",
            )
        )

    return pos, lines, issues


# --------------------------------------------------------------------------
CONTACTS = [
    dict(name="Danny (Daniel Augustyn)", role="Owner / CEO — RepLabs", org="RepLabs",
         provides="Product direction, new SKU decisions, PO authorisation", channel="Text / in person"),
    dict(name="Lowry Akers", role="VP Operations — Powder Ops", org="Powder Ops",
         provides="Owns this system", channel="—"),
    dict(name="Shaun", role="Designer", org="External / Sipo Studios",
         provides="Artwork, dielines, proofs, NFP layout", channel="Text"),
    dict(name="Matt Schramm", role="Formulator / R&D — M4 Dynamic", org="M4 Dynamic (contract)",
         provides="Formulas, NFP data, ingredient/allergen statements", channel="Text / in person"),
    dict(name="Michael Rino", role="Account contact", org="Professional Packaging Services (PPS)",
         provides="Quotes, PO confirmations, proofs, lead times, ship dates", channel="Email / phone"),
    dict(name="Adam Bliss", role="Production Manager", org="Powder Ops",
         provides="Run feasibility, material consumption, defects", channel="In person"),
    dict(name="Jake Waits", role="Procurement & Inventory Manager", org="Powder Ops",
         provides="Receipts, on-hand counts, reorder triggers", channel="In person"),
]

CHANGE_LOG_SEED = [
    dict(id="CHG-0001", date_logged="2026-07-29", source="Lowry (system audit)", channel="Audit",
         area="PO", sku_affected="", raw_note="Film PO and Pancake PO both carry lines annotated as removed, but the annotated lines are still inside the SUM. Combined overstatement 140,000 units / $24,850.",
         decision="Confirm true quantities with PPS before invoicing", owner="Lowry",
         due="2026-08-01", status="Open", priority="Critical"),
    dict(id="CHG-0002", date_logged="2026-07-29", source="Lowry (system audit)", channel="Audit",
         area="PO", sku_affected="", raw_note="Two of three PO file names disagree with the PO number inside the sheet (41626 vs 42826, 42926 vs 41726).",
         decision="Adopt auto-generated PO numbers; never hand-type", owner="Lowry",
         due="2026-08-01", status="Open", priority="High"),
    dict(id="CHG-0003", date_logged="2026-07-29", source="Lowry (system audit)", channel="Audit",
         area="SKU", sku_affected="PP-PA-30", raw_note="Pomegranate Acai has stick film on the film PO (55,000) but no pouch on the pouch PO.",
         decision="Confirm whether the pouch is intentionally deferred", owner="Lowry",
         due="2026-08-05", status="Open", priority="High"),
    dict(id="CHG-0004", date_logged="2026-07-29", source="Lowry (system audit)", channel="Audit",
         area="Artwork", sku_affected="PP-CM-02, PSP-CM, PP-RC-12, PSP-RC, PP-BF-21, PSP-BF",
         raw_note="Six SKUs carry hex values that are not valid hex (4E2CID, F43BF, E613B24).",
         decision="Shaun to supply corrected values from the print files", owner="Shaun",
         due="2026-08-05", status="Open", priority="High"),
    dict(id="CHG-0005", date_logged="2026-07-29", source="Lowry (system audit)", channel="Audit",
         area="SKU", sku_affected="Beef protein pouches (4)",
         raw_note="Four beef pouch SKUs hold Shopify variant IDs instead of SKU codes.",
         decision="Assign real SKU codes matching the HBF-* stick convention", owner="Lowry",
         due="2026-08-05", status="Open", priority="High"),
]


def main():
    os.makedirs(OUT, exist_ok=True)
    print("Building normalized dataset...")

    products, colors, issues_a = build_products()
    pos, po_lines, issues_b = build_pos(products)
    issues = issues_a + issues_b

    order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    issues.sort(key=lambda i: (order[i["severity"]], i["area"], i["sku"]))
    for n, i in enumerate(issues, 1):
        i["issue_id"] = f"ISS-{n:04d}"

    write("products.csv", list(products[0].keys()), products)
    write("packaging_specs.csv", list(SPECS[0].keys()), SPECS)
    write("sku_colors.csv", ["sku", "slot", "pms", "hex", "hex_valid", "pms_valid"], colors)
    write("purchase_orders.csv", list(pos[0].keys()), pos)
    write("po_lines.csv", list(po_lines[0].keys()), po_lines)
    write("contacts.csv", list(CONTACTS[0].keys()), CONTACTS)
    write("change_log.csv", list(CHANGE_LOG_SEED[0].keys()), CHANGE_LOG_SEED)
    write("data_issues.csv",
          ["issue_id", "severity", "area", "sku", "flavor", "issue", "detail"], issues)

    write("artwork_versions.csv",
          ["artwork_id", "sku", "version", "date", "designer", "stage", "change_summary",
           "proof_link", "print_file_link", "gtin_verified", "nfp_verified",
           "eyemark_verified", "approved_by", "approved_date", "sent_to_vendor_date"],
          [])

    print("\nSummary")
    print(f"  products            {len(products)}")
    print(f"  packaging specs     {len(SPECS)}")
    print(f"  colour slots        {len(colors)}")
    print(f"  purchase orders     {len(pos)}")
    print(f"  PO lines            {len(po_lines)}")
    print(f"  open data issues    {len(issues)}")
    for sev in ("Critical", "High", "Medium", "Low"):
        c = sum(1 for i in issues if i["severity"] == sev)
        print(f"    {sev:<9} {c}")
    print("\n  PO totals as written vs corrected:")
    for p in pos:
        print(f"    PO {p['po_number']:<8} {p['total_qty_as_written']:>10,} u / "
              f"${p['total_cost_as_written']:>11,.2f}   ->   "
              f"{p['total_qty_corrected']:>10,} u / ${p['total_cost_corrected']:>11,.2f}")


if __name__ == "__main__":
    main()
